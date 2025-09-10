#!/usr/bin/env python3
"""
Excel Export Module for E069: Streamlined Scenario Batch Processing

Export simulation results to analyst-friendly Excel workbooks with metadata and splitting options.

Features:
- Multi-year workforce snapshot export with optional per-year sheet splitting
- Summary metrics sheet with key workforce KPIs
- Event counts sheet with hire/termination/promotion trends
- Metadata sheet with scenario config, git SHA, pipeline version, random seed
- Professional formatting with auto-sized columns
- CSV fallback support
- Comparison workbook generation across scenarios
"""

from __future__ import annotations

import json
import subprocess
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

from .utils import DatabaseConnectionManager


class ExcelExporter:
    """Export simulation results to analyst-friendly Excel workbooks with metadata and splitting options."""

    def __init__(self, db_manager: DatabaseConnectionManager, *, split_threshold: int = 750_000):
        """Initialize Excel exporter with database connection manager.

        Args:
            db_manager: Database connection manager for accessing simulation data
            split_threshold: Row threshold for splitting workforce snapshot into per-year sheets
        """
        self.db_manager = db_manager
        self.split_threshold = split_threshold

    def export_scenario_results(
        self,
        *,
        scenario_name: str,
        output_dir: Path,
        config: Any,
        seed: int,
        export_format: str = "excel",
        split_by_year: Optional[bool] = None,
    ) -> Path:
        """Export complete scenario results to Excel workbook or CSV files.

        Args:
            scenario_name: Name of the scenario being exported
            output_dir: Output directory for exported files
            config: SimulationConfig object with scenario parameters
            seed: Random seed used for the simulation
            export_format: Export format ('excel' or 'csv')
            split_by_year: Force splitting by year (auto-determined if None)

        Returns:
            Path to the main export file or directory
        """
        conn = self.db_manager.get_connection()
        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            # Check if workforce snapshot table exists
            table_exists = self._check_table_exists(conn, "fct_workforce_snapshot")
            if not table_exists:
                print(f"   ⚠️  Warning: fct_workforce_snapshot table not found, creating minimal export")
                return self._create_minimal_export(scenario_name, output_dir, export_format)

            # Determine total rows and whether to split by year
            total_rows = pd.read_sql("SELECT COUNT(*) AS cnt FROM fct_workforce_snapshot", conn)["cnt"].iloc[0]
            split = split_by_year if split_by_year is not None else total_rows > self.split_threshold

            if export_format.lower() == "csv":
                return self._export_csv(scenario_name, output_dir, conn, config, seed, split)
            else:
                return self._export_excel(scenario_name, output_dir, conn, config, seed, split, total_rows)

        finally:
            conn.close()

    def _check_table_exists(self, conn, table_name: str) -> bool:
        """Check if a table exists in the database.

        Args:
            conn: Database connection
            table_name: Name of the table to check

        Returns:
            True if table exists, False otherwise
        """
        try:
            result = conn.execute(f"SELECT COUNT(*) FROM information_schema.tables WHERE table_name = '{table_name}'").fetchone()
            return result[0] > 0 if result else False
        except Exception:
            # Try direct query as fallback
            try:
                conn.execute(f"SELECT COUNT(*) FROM {table_name} LIMIT 1").fetchone()
                return True
            except Exception:
                return False

    def _create_minimal_export(self, scenario_name: str, output_dir: Path, export_format: str) -> Path:
        """Create minimal export when main tables are not available.

        Args:
            scenario_name: Name of the scenario
            output_dir: Output directory
            export_format: Export format

        Returns:
            Path to the minimal export file
        """
        if export_format.lower() == "csv":
            minimal_path = output_dir / f"{scenario_name}_minimal_export.csv"
            pd.DataFrame({"message": ["Simulation completed but no data tables found"]}).to_csv(
                minimal_path, index=False
            )
            return minimal_path
        else:
            minimal_path = output_dir / f"{scenario_name}_minimal_export.xlsx"
            with pd.ExcelWriter(minimal_path, engine="openpyxl") as writer:
                pd.DataFrame({"message": ["Simulation completed but no data tables found"]}).to_excel(
                    writer, sheet_name="Status", index=False
                )
            return minimal_path

    def _export_csv(self, scenario_name: str, output_dir: Path, conn, config: Any, seed: int, split: bool) -> Path:
        """Export scenario results to CSV files.

        Args:
            scenario_name: Name of the scenario
            output_dir: Output directory
            conn: Database connection
            config: SimulationConfig object
            seed: Random seed
            split: Whether to split workforce snapshot by year

        Returns:
            Path to the output directory containing CSV files
        """
        # Export workforce snapshot
        if split:
            # Export per-year CSV files
            years = pd.read_sql(
                "SELECT DISTINCT simulation_year FROM fct_workforce_snapshot ORDER BY simulation_year",
                conn,
            )["simulation_year"].tolist()

            for year in years:
                df_year = pd.read_sql(
                    "SELECT * FROM fct_workforce_snapshot WHERE simulation_year = ? ORDER BY employee_id",
                    conn,
                    params=[year],
                )
                df_year.to_csv(output_dir / f"{scenario_name}_workforce_{year}.csv", index=False)
        else:
            # Single workforce snapshot CSV
            df_workforce = pd.read_sql(
                "SELECT * FROM fct_workforce_snapshot ORDER BY simulation_year, employee_id",
                conn,
            )
            df_workforce.to_csv(output_dir / f"{scenario_name}_workforce_snapshot.csv", index=False)

        # Export summary metrics
        df_summary = self._calculate_summary_metrics(conn)
        df_summary.to_csv(output_dir / f"{scenario_name}_summary_metrics.csv", index=False)

        # Export events summary if table exists
        if self._check_table_exists(conn, "fct_yearly_events"):
            df_events = self._calculate_events_summary(conn)
            df_events.to_csv(output_dir / f"{scenario_name}_events_summary.csv", index=False)

        # Export metadata
        df_metadata = self._build_metadata_dataframe(config, seed, conn, split=split)
        df_metadata.to_csv(output_dir / f"{scenario_name}_metadata.csv", index=False)

        return output_dir

    def _export_excel(self, scenario_name: str, output_dir: Path, conn, config: Any, seed: int, split: bool, total_rows: int) -> Path:
        """Export scenario results to Excel workbook.

        Args:
            scenario_name: Name of the scenario
            output_dir: Output directory
            conn: Database connection
            config: SimulationConfig object
            seed: Random seed
            split: Whether to split workforce snapshot by year
            total_rows: Total number of rows in workforce snapshot

        Returns:
            Path to the Excel workbook
        """
        excel_path = output_dir / f"{scenario_name}_results.xlsx"

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # Workforce Snapshot (single sheet or per-year)
            self._write_workforce_sheets(writer, conn, split)

            # Summary Metrics by Year
            df_summary = self._calculate_summary_metrics(conn)
            df_summary.to_excel(writer, sheet_name="Summary_Metrics", index=False)
            self._format_worksheet(writer.book["Summary_Metrics"])

            # Events Summary (if events table exists)
            if self._check_table_exists(conn, "fct_yearly_events"):
                df_events = self._calculate_events_summary(conn)
                df_events.to_excel(writer, sheet_name="Events_Summary", index=False)
                self._format_worksheet(writer.book["Events_Summary"])

            # Metadata sheet
            df_metadata = self._build_metadata_dataframe(config, seed, conn, total_rows, split)
            df_metadata.to_excel(writer, sheet_name="Metadata", index=False)
            self._format_worksheet(writer.book["Metadata"])

        return excel_path

    def _write_workforce_sheets(self, writer, conn, split: bool) -> None:
        """Write workforce snapshot data to Excel sheets.

        Args:
            writer: Excel writer object
            conn: Database connection
            split: Whether to split by year
        """
        if split:
            # Create per-year sheets
            years = pd.read_sql(
                "SELECT DISTINCT simulation_year FROM fct_workforce_snapshot ORDER BY simulation_year",
                conn,
            )["simulation_year"].tolist()

            for year in years:
                df_year = pd.read_sql(
                    "SELECT * FROM fct_workforce_snapshot WHERE simulation_year = ? ORDER BY employee_id",
                    conn,
                    params=[year],
                )
                sheet_name = f"Workforce_{year}"
                df_year.to_excel(writer, sheet_name=sheet_name, index=False)
                self._format_worksheet(writer.book[sheet_name])
        else:
            # Single workforce snapshot sheet
            df_workforce = pd.read_sql(
                "SELECT * FROM fct_workforce_snapshot ORDER BY simulation_year, employee_id",
                conn,
            )
            df_workforce.to_excel(writer, sheet_name="Workforce_Snapshot", index=False)
            self._format_worksheet(writer.book["Workforce_Snapshot"])

    def _calculate_summary_metrics(self, conn) -> pd.DataFrame:
        """Calculate summary metrics by simulation year.

        Args:
            conn: Database connection

        Returns:
            DataFrame with summary metrics by year
        """
        return pd.read_sql(
            """
            SELECT
                simulation_year,
                COUNT(*) AS total_employees,
                COUNT(CASE WHEN status = 'active' THEN 1 END) AS active_employees,
                COUNT(CASE WHEN enrollment_date IS NOT NULL THEN 1 END) AS enrolled_employees,
                ROUND(AVG(current_salary), 2) AS avg_salary,
                ROUND(SUM(COALESCE(employee_contribution_annual, 0)), 2) AS total_employee_contributions,
                ROUND(SUM(COALESCE(employer_match_annual, 0)), 2) AS total_employer_match,
                ROUND(AVG(COALESCE(deferral_rate, 0)), 4) AS avg_deferral_rate
            FROM fct_workforce_snapshot
            GROUP BY simulation_year
            ORDER BY simulation_year
            """,
            conn,
        )

    def _calculate_events_summary(self, conn) -> pd.DataFrame:
        """Calculate events summary by simulation year and event type.

        Args:
            conn: Database connection

        Returns:
            DataFrame with event counts and metrics
        """
        return pd.read_sql(
            """
            SELECT
                simulation_year,
                event_type,
                COUNT(*) AS event_count,
                ROUND(AVG(CASE
                    WHEN event_type = 'raise' AND json_extract_scalar(event_data, '$.new_salary') IS NOT NULL
                    THEN CAST(json_extract_scalar(event_data, '$.new_salary') AS DOUBLE)
                    END), 2) AS avg_new_salary_on_raise
            FROM fct_yearly_events
            GROUP BY simulation_year, event_type
            ORDER BY simulation_year, event_type
            """,
            conn,
        )

    def _build_metadata_dataframe(self, config: Any, seed: int, conn, total_rows: Optional[int] = None, split: bool = False) -> pd.DataFrame:
        """Build metadata dataframe with scenario information.

        Args:
            config: SimulationConfig object
            seed: Random seed used
            conn: Database connection
            total_rows: Total rows in workforce snapshot (optional)
            split: Whether snapshot was split by year

        Returns:
            DataFrame with metadata information
        """
        # Get git metadata
        git_metadata = self._get_git_metadata()

        # Get simulation year range
        try:
            years = pd.read_sql(
                "SELECT MIN(simulation_year) AS min_y, MAX(simulation_year) AS max_y FROM fct_workforce_snapshot",
                conn,
            ).iloc[0]
            start_year = int(years["min_y"]) if not pd.isna(years["min_y"]) else config.simulation.start_year
            end_year = int(years["max_y"]) if not pd.isna(years["max_y"]) else config.simulation.end_year
        except Exception:
            start_year = config.simulation.start_year
            end_year = config.simulation.end_year

        # Get total rows if not provided
        if total_rows is None:
            try:
                total_rows = pd.read_sql("SELECT COUNT(*) AS cnt FROM fct_workforce_snapshot", conn)["cnt"].iloc[0]
            except Exception:
                total_rows = 0

        # Convert config to JSON-serializable dictionary
        try:
            config_dict = config.model_dump() if hasattr(config, 'model_dump') else {}
        except Exception:
            config_dict = {}

        metadata_record = {
            "export_timestamp": pd.Timestamp.now().isoformat(),
            "git_sha": git_metadata.get("git_sha", "unknown"),
            "git_branch": git_metadata.get("git_branch", "unknown"),
            "git_clean": git_metadata.get("git_clean", False),
            "random_seed": seed,
            "start_year": start_year,
            "end_year": end_year,
            "workforce_rows": int(total_rows),
            "snapshot_split_by_year": bool(split),
            "target_growth_rate": config.simulation.target_growth_rate,
            "cola_rate": config.compensation.cola_rate,
            "merit_budget": config.compensation.merit_budget,
            "config_json": json.dumps(config_dict, default=str, indent=2)
        }

        # Convert to single-row DataFrame with transposed layout for readability
        metadata_items = []
        for key, value in metadata_record.items():
            if key == "config_json":
                # Split config JSON into multiple rows for readability
                try:
                    config_lines = str(value).split('\n')
                    for i, line in enumerate(config_lines):
                        metadata_items.append({
                            "Parameter": f"config_json_line_{i+1:03d}",
                            "Value": line
                        })
                except Exception:
                    metadata_items.append({"Parameter": key, "Value": str(value)})
            else:
                metadata_items.append({"Parameter": key, "Value": str(value)})

        return pd.DataFrame(metadata_items)

    def _get_git_metadata(self) -> Dict[str, Any]:
        """Get git metadata for the current repository state.

        Returns:
            Dictionary with git metadata (SHA, branch, etc.)
        """
        metadata = {}

        try:
            # Get git SHA
            result = subprocess.run(
                ["git", "rev-parse", "HEAD"],
                capture_output=True,
                text=True,
                check=True
            )
            metadata["git_sha"] = result.stdout.strip()
        except Exception:
            metadata["git_sha"] = "unknown"

        try:
            # Get git branch
            result = subprocess.run(
                ["git", "rev-parse", "--abbrev-ref", "HEAD"],
                capture_output=True,
                text=True,
                check=True
            )
            metadata["git_branch"] = result.stdout.strip()
        except Exception:
            metadata["git_branch"] = "unknown"

        try:
            # Check if working directory is clean
            result = subprocess.run(
                ["git", "status", "--porcelain"],
                capture_output=True,
                text=True,
                check=True
            )
            metadata["git_clean"] = len(result.stdout.strip()) == 0
        except Exception:
            metadata["git_clean"] = False

        return metadata

    def _format_worksheet(self, worksheet) -> None:
        """Apply professional formatting to an Excel worksheet.

        Args:
            worksheet: openpyxl Worksheet object to format
        """
        try:
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            # Header formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            # Format header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Auto-size columns
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                column_letter = get_column_letter(column_cells[0].column)
                # Set reasonable max width
                worksheet.column_dimensions[column_letter].width = min(length + 2, 50)

            # Freeze header row
            worksheet.freeze_panes = "A2"

        except ImportError:
            # openpyxl not available, skip formatting
            pass
        except Exception:
            # Formatting failed, continue without formatting
            pass

    def create_comparison_workbook(self, scenario_results: Dict[str, Any], output_path: Path) -> None:
        """Create comparison workbook across multiple successful scenarios.

        Args:
            scenario_results: Dictionary mapping scenario names to their results
            output_path: Path for the comparison workbook
        """
        try:
            comparison_data = []

            for scenario_name, result in scenario_results.items():
                db_path = Path(result["database_path"])
                if not db_path.exists():
                    continue

                # Connect to scenario database and extract summary metrics
                scenario_db = DatabaseConnectionManager(db_path)
                conn = scenario_db.get_connection()

                try:
                    # Get summary metrics for this scenario
                    summary_df = self._calculate_summary_metrics(conn)

                    for _, row in summary_df.iterrows():
                        comparison_record = {
                            "scenario": scenario_name,
                            "simulation_year": row["simulation_year"],
                            "total_employees": row["total_employees"],
                            "active_employees": row["active_employees"],
                            "enrolled_employees": row["enrolled_employees"],
                            "avg_salary": row["avg_salary"],
                            "total_employee_contributions": row["total_employee_contributions"],
                            "total_employer_match": row["total_employer_match"],
                            "avg_deferral_rate": row["avg_deferral_rate"],
                            "execution_time_seconds": result.get("execution_time_seconds", 0),
                            "seed": result.get("seed", 0)
                        }
                        comparison_data.append(comparison_record)

                finally:
                    conn.close()

            if comparison_data:
                # Create comparison workbook
                comparison_df = pd.DataFrame(comparison_data)

                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    # Main comparison sheet
                    comparison_df.to_excel(writer, sheet_name="Scenario_Comparison", index=False)
                    self._format_worksheet(writer.book["Scenario_Comparison"])

                    # Pivot table by scenario and year
                    pivot_df = comparison_df.pivot_table(
                        index=["simulation_year"],
                        columns=["scenario"],
                        values=["total_employees", "avg_salary", "avg_deferral_rate"],
                        aggfunc="first"
                    )
                    pivot_df.to_excel(writer, sheet_name="Year_by_Scenario_Pivot")

                    # Summary statistics
                    summary_stats = comparison_df.groupby("scenario").agg({
                        "total_employees": ["min", "max", "mean"],
                        "avg_salary": ["min", "max", "mean"],
                        "avg_deferral_rate": ["min", "max", "mean"],
                        "execution_time_seconds": "first"
                    }).round(2)
                    summary_stats.to_excel(writer, sheet_name="Scenario_Summary")

                print(f"   📊 Comparison workbook created with {len(comparison_data)} data points")
            else:
                print(f"   ⚠️  No comparison data found for scenarios")

        except Exception as e:
            print(f"   ❌ Error creating comparison workbook: {e}")
            raise
