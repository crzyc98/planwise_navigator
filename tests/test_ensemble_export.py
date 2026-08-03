"""Workbook coverage for optional seed-ensemble distribution exports."""

from __future__ import annotations

from types import SimpleNamespace

import duckdb
import pytest
from openpyxl import load_workbook

from planalign_cli.commands.batch import _export_ensemble_workbook
from planalign_orchestrator.excel_exporter import ExcelExporter
from planalign_orchestrator.utils import DatabaseConnectionManager


def _config() -> SimpleNamespace:
    """Supply only the ordinary-export metadata fields required by this fixture."""
    return SimpleNamespace(
        simulation=SimpleNamespace(
            start_year=2029,
            end_year=2029,
            random_seed=42,
            target_growth_rate=0.03,
        ),
        compensation=SimpleNamespace(cola_rate=0.02, merit_budget=0.03),
    )


def _write_primary_snapshot(path) -> None:
    """Create a minimal normal result database for the existing workbook sheets."""
    with duckdb.connect(str(path)) as conn:
        conn.execute(
            "CREATE TABLE fct_workforce_snapshot ("
            "employee_id VARCHAR, simulation_year INTEGER, employment_status VARCHAR)"
        )
        conn.execute(
            "INSERT INTO fct_workforce_snapshot VALUES ('employee-1', 2029, 'active')"
        )


def _write_ensemble_distributions(path) -> None:
    """Create sufficient and withheld aggregate rows in a dedicated DB."""
    with duckdb.connect(str(path)) as conn:
        conn.execute(
            "CREATE TABLE fct_metric_distributions ("
            "metric VARCHAR, simulation_year INTEGER, p10 DOUBLE, p25 DOUBLE, "
            "p50 DOUBLE, p75 DOUBLE, p90 DOUBLE, mean DOUBLE, stddev DOUBLE, "
            "n_seeds INTEGER, n_seeds_requested INTEGER, is_sufficient BOOLEAN, "
            "percentile_method VARCHAR)"
        )
        conn.execute(
            "INSERT INTO fct_metric_distributions VALUES "
            "('total_employer_plan_cost', 2029, 1.0, 2.0, 3.0, 4.0, 5.0, 3.0, "
            "1.0, 10, 10, TRUE, 'linear'), "
            "('participation_rate', 2029, NULL, NULL, NULL, NULL, NULL, NULL, "
            "NULL, 4, 10, FALSE, 'linear')"
        )


def _write_ensemble_attribution(path) -> None:
    """Create one measured and one structural attribution row for export."""
    with duckdb.connect(str(path)) as conn:
        conn.execute(
            "CREATE TABLE fct_variance_attribution ("
            "metric VARCHAR, simulation_year INTEGER, subsystem VARCHAR, "
            "variance_share DOUBLE, baseline_variance DOUBLE, frozen_variance DOUBLE, "
            "n_seeds INTEGER, baselines_reused INTEGER, baselines_executed INTEGER, "
            "stochastic_status VARCHAR)"
        )
        conn.execute(
            "INSERT INTO fct_variance_attribution VALUES "
            "('total_employer_plan_cost', 2029, 'termination', 0.61, 10, 3.9, "
            "10, 10, 0, 'stochastic'), "
            "('total_employer_plan_cost', 2029, 'enrollment', NULL, NULL, NULL, "
            "10, 10, 0, 'not_stochastic')"
        )


@pytest.mark.fast
def test_distribution_sheet_matches_stored_rows_and_preserves_nulls(tmp_path) -> None:
    """Withheld bands become empty cells rather than misleading zeroes."""
    primary_db = tmp_path / "seed.duckdb"
    ensemble_db = tmp_path / "ensemble.duckdb"
    _write_primary_snapshot(primary_db)
    _write_ensemble_distributions(ensemble_db)
    manager = DatabaseConnectionManager(primary_db)
    exporter = ExcelExporter(manager)
    exporter._get_git_metadata = lambda: {
        "git_sha": "test",
        "git_branch": "test",
        "git_clean": True,
    }

    try:
        workbook_path = exporter.export_scenario_results(
            scenario_name="baseline",
            output_dir=tmp_path,
            config=_config(),
            seed=42,
            ensemble_db_path=ensemble_db,
        )
    finally:
        manager.close_all()

    workbook = load_workbook(workbook_path, data_only=True)
    assert "Metric_Distributions" in workbook.sheetnames
    sheet = workbook["Metric_Distributions"]
    header = [cell.value for cell in sheet[1]]
    rows = [
        dict(zip(header, row, strict=True))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]
    by_metric = {row["metric"]: row for row in rows}
    assert by_metric["total_employer_plan_cost"]["p50"] == 3.0
    assert by_metric["participation_rate"]["p10"] is None
    assert by_metric["participation_rate"]["is_sufficient"] is False


@pytest.mark.fast
def test_normal_export_does_not_gain_an_empty_distribution_sheet(tmp_path) -> None:
    """Existing non-ensemble workbooks remain structurally unchanged."""
    primary_db = tmp_path / "seed.duckdb"
    _write_primary_snapshot(primary_db)
    manager = DatabaseConnectionManager(primary_db)
    exporter = ExcelExporter(manager)
    exporter._get_git_metadata = lambda: {
        "git_sha": "test",
        "git_branch": "test",
        "git_clean": True,
    }

    try:
        workbook_path = exporter.export_scenario_results(
            scenario_name="baseline", output_dir=tmp_path, config=_config(), seed=42
        )
    finally:
        manager.close_all()

    sheetnames = load_workbook(workbook_path).sheetnames
    assert "Metric_Distributions" not in sheetnames
    assert "Variance_Attribution" not in sheetnames


@pytest.mark.fast
def test_attribution_is_withheld_from_the_client_workbook_but_kept_as_evidence(
    tmp_path,
) -> None:
    """A spreadsheet strips the caveats single-anchor variance cannot be read without."""
    primary_db = tmp_path / "seed.duckdb"
    ensemble_db = tmp_path / "ensemble.duckdb"
    _write_primary_snapshot(primary_db)
    _write_ensemble_distributions(ensemble_db)
    _write_ensemble_attribution(ensemble_db)
    manager = DatabaseConnectionManager(primary_db)
    exporter = ExcelExporter(manager)
    exporter._get_git_metadata = lambda: {
        "git_sha": "test",
        "git_branch": "test",
        "git_clean": True,
    }

    try:
        workbook_path = exporter.export_scenario_results(
            scenario_name="baseline",
            output_dir=tmp_path,
            config=_config(),
            seed=42,
            ensemble_db_path=ensemble_db,
        )
    finally:
        manager.close_all()

    sheetnames = load_workbook(workbook_path, data_only=True).sheetnames
    assert "Variance_Attribution" not in sheetnames
    assert "Metric_Distributions" in sheetnames

    # Withheld from the workbook, never deleted: the evidence stays queryable.
    with duckdb.connect(str(ensemble_db), read_only=True) as connection:
        retained = connection.execute(
            "SELECT subsystem, variance_share FROM fct_variance_attribution "
            "ORDER BY subsystem"
        ).fetchall()
    assert dict(retained)["termination"] == 0.61


@pytest.mark.fast
def test_aggregate_only_export_does_not_recreate_discarded_seed_db(tmp_path) -> None:
    """A completed aggregate remains exportable after per-seed cleanup."""
    primary_db = tmp_path / "discarded_seed.duckdb"
    ensemble_db = tmp_path / "ensemble.duckdb"
    _write_ensemble_distributions(ensemble_db)
    manager = DatabaseConnectionManager(primary_db)

    try:
        workbook_path = ExcelExporter(manager).export_scenario_results(
            scenario_name="baseline",
            output_dir=tmp_path,
            config=_config(),
            seed=42,
            ensemble_db_path=ensemble_db,
        )
    finally:
        manager.close_all()

    assert not primary_db.exists()
    assert "Metric_Distributions" in load_workbook(workbook_path).sheetnames


@pytest.mark.fast
def test_batch_ensemble_workbook_reads_the_dedicated_aggregate(tmp_path) -> None:
    """Batch export can hand off bands without choosing a representative seed."""
    ensemble_db = tmp_path / "ensemble.duckdb"
    _write_ensemble_distributions(ensemble_db)
    batch_runner = SimpleNamespace(batch_output_dir=tmp_path / "batch")
    result = SimpleNamespace(plan=SimpleNamespace(ensemble_db_path=ensemble_db))

    workbook_path = _export_ensemble_workbook(
        batch_runner=batch_runner,
        scenario_name="baseline",
        config=_config(),
        result=result,
    )

    assert workbook_path == tmp_path / "batch" / "baseline" / "baseline_results.xlsx"
    assert "Metric_Distributions" in load_workbook(workbook_path).sheetnames
